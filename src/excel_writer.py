"""추출된 표를 원본 구조(병합/크기/제목)를 유지하며 엑셀로 저장."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from .models import MatchResult, Table

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

TABLE_START_ROW_WITH_TITLE = 3  # 1행: 제목, 2행: 여백, 3행부터 표
TABLE_START_ROW_NO_TITLE = 1    # 제목 없이 표만 → 1행부터
TABLE_GAP_ROWS = 2              # 연속 배치 시 표 블록 사이 빈 행 수

SHEET_LAYOUT_PER_TABLE = "per_table"
SHEET_LAYOUT_SINGLE = "single_sheet"


def hwpunit_to_pt(v: int) -> float:
    """HWPUNIT(1/7200 inch) → 포인트(1/72 inch). 행 높이용."""
    return v / 100


def hwpunit_to_col_width(v: int) -> float:
    """HWPUNIT → 엑셀 열 너비(문자 폭 단위) 근사 변환.

    1) HWPUNIT → 인치 → 96dpi 픽셀
    2) 엑셀 열 너비 = (픽셀 - 5) / 7  (기본 폰트 기준 근사식)
    """
    px = v / 7200 * 96
    return max((px - 5) / 7, 2.0)


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """엑셀 시트명 제약(31자, 금지문자) 처리 및 중복 방지."""
    name = re.sub(r"[\\/*?:\[\]]", "_", name).strip() or "표"
    name = name[:28]
    candidate, n = name, 1
    while candidate in used:
        n += 1
        candidate = f"{name}_{n}"
    used.add(candidate)
    return candidate


def _resolve_block_start(start_row: int | None, *, include_title: bool) -> tuple[int, int]:
    """(블록 시작 행, 표 본문 시작 행) 반환."""
    if start_row is None:
        block_start = 1
        table_start = TABLE_START_ROW_WITH_TITLE if include_title else TABLE_START_ROW_NO_TITLE
        return block_start, table_start
    table_start = start_row + 2 if include_title else start_row
    return start_row, table_start


def _set_column_width(ws, col: int, width_hwpunit: int) -> None:
    letter = get_column_letter(col)
    new_width = hwpunit_to_col_width(width_hwpunit)
    current = ws.column_dimensions[letter].width
    if current is None or new_width > current:
        ws.column_dimensions[letter].width = new_width


def write_table(
    ws,
    table: Table,
    *,
    start_row: int | None = None,
    include_title: bool = True,
    include_footer: bool = True,
) -> int:
    """표(및 꼬리말)를 기록하고 마지막으로 사용한 행 번호를 반환."""
    block_start, table_start = _resolve_block_start(start_row, include_title=include_title)

    if include_title:
        title = table.title()
        if title:
            tc = ws.cell(row=block_start, column=1, value=title)
            tc.font = Font(bold=True, size=12)

    row_heights: dict[int, int] = {}

    for c in table.cells:
        r = table_start + c.row
        col = 1 + c.col

        cell = ws.cell(row=r, column=col, value=c.text if c.text else None)
        cell.border = _BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center")

        if c.row_span > 1 or c.col_span > 1:
            end_r, end_c = r + c.row_span - 1, col + c.col_span - 1
            ws.merge_cells(start_row=r, start_column=col, end_row=end_r, end_column=end_c)
            for rr in range(r, end_r + 1):
                for cc in range(col, end_c + 1):
                    ws.cell(row=rr, column=cc).border = _BORDER

        if c.col_span == 1 and c.width > 0:
            _set_column_width(ws, col, c.width)
        if c.row_span == 1 and c.height > 0:
            row_heights[r] = max(row_heights.get(r, 0), c.height)

    for r, h in row_heights.items():
        ws.row_dimensions[r].height = hwpunit_to_pt(h)

    last_row = table_start + table.n_rows - 1

    footer = table.footer_text.strip() if include_footer else ""
    if footer:
        footer_row = table_start + table.n_rows
        footer_cell = ws.cell(row=footer_row, column=1, value=footer)
        footer_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if table.n_cols > 1:
            ws.merge_cells(
                start_row=footer_row,
                start_column=1,
                end_row=footer_row,
                end_column=table.n_cols,
            )
        last_row = footer_row

    return last_row


def write_workbook(
    results: list[MatchResult],
    out_path: str | Path,
    *,
    include_title: bool = True,
    include_footer: bool = True,
    sheet_layout: str = SHEET_LAYOUT_PER_TABLE,
) -> Path:
    """매칭된 표들을 엑셀 파일로 저장."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    if not results:
        ws = wb.create_sheet("결과없음")
        ws.cell(row=1, column=1, value="선별 조건에 맞는 표를 찾지 못했습니다.")
    elif sheet_layout == SHEET_LAYOUT_SINGLE:
        ws = wb.create_sheet("추출표")
        next_row = 1
        for i, res in enumerate(results):
            last_row = write_table(
                ws,
                res.table,
                start_row=next_row,
                include_title=include_title,
                include_footer=include_footer,
            )
            if i < len(results) - 1:
                next_row = last_row + 1 + TABLE_GAP_ROWS
    else:
        used_names: set[str] = set()
        for res in results:
            sheet_name = _safe_sheet_name(f"{res.rule_name}_{res.table.index}", used_names)
            ws = wb.create_sheet(sheet_name)
            write_table(
                ws,
                res.table,
                include_title=include_title,
                include_footer=include_footer,
            )

    wb.save(out_path)
    return out_path
